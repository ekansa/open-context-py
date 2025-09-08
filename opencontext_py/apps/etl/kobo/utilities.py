from copy import copy
import json
import os
import numpy as np
import pandas as pd
from openpyxl import load_workbook
import re
import requests
import openpyxl

from django.db.models import Q
from opencontext_py.apps.all_items.models import AllManifest

from opencontext_py.apps.etl.kobo import pc_configs


MULTI_VALUE_COL_PREFIXES = [
    'Preliminary Phasing/',
    'Trench Supervisor/',
    'Decorative Techniques and Motifs/Decorative Technique/',
    'Decorative Techniques and Motifs/Motif/',
    'Fabric Category/',
    'Vessel Part Present/',
    'Modification/',
    'Type of Composition Subject/',
    'Object General Type Alternatives/',
    'Object Type/',
    'Recovery Techniques/',
    'Locus Period/',
]

UUID_SOURCE_KOBOTOOLBOX = 'kobotoolbox' # UUID minted by kobotoolbox
UUID_SOURCE_OC_KOBO_ETL = 'oc-kobo-etl' # UUID minted by this ETL process
UUID_SOURCE_OC_LOOKUP = 'open-context' # UUID existing in Open Context

LINK_RELATION_TYPE_COL = 'relation_type'

SKIP_UNDERSCORE_FIX_COLS = [
    'subject_label',
    'subject_uuid',
    'subject_uuid_source',
    'trench_id',
    'trench_year',
    '_id',
    '_uuid',
    '_submission_time',
    '_validation_status',
    '_notes',
    '_status',
    '_submitted_by',
    '__version__',
    '_tags',
    '_index',
    '_submission__uuid',
    '_orig_uuid',
    '_orig_submission__uuid',
]



def fix_df_col_underscores(df):
    """Fixes underscores introduced in the 2024 data export, because of course"""
    if pc_configs.DEFAULT_IMPORT_YEAR != 2024:
        return df
    r_cols = {c:c.replace('_', ' ') for c in df.columns.tolist() if c not in SKIP_UNDERSCORE_FIX_COLS}
    df.rename(columns=r_cols, inplace=True)
    # print(f'Removed underscores from {len(r_cols)} columns')
    return df



def make_directory_files_df(attachments_path):
    """Makes a dataframe listing all the files a Kobo Attachments directory."""
    file_data = []
    for dirpath, _, filenames in os.walk(attachments_path):
        for filename in filenames:
            if filename.endswith(':Zone.Identifier'):
                # A convenience hack for Windows subsystem for linux
                continue
            file_path = os.path.join(dirpath, filename)
            uuid_dir = os.path.split(os.path.abspath(dirpath))[-1]
            rec = {
                'path': file_path,
                'path_uuid': uuid_dir,
                'filename': filename,
            }
            file_data.append(rec)
    df = pd.DataFrame(data=file_data)
    return df

def list_excel_files(excel_dirpath):
    """Makes a list of Excel files in a directory path."""
    xlsx_files = []
    for item in os.listdir(excel_dirpath):
        item_path = os.path.join(excel_dirpath, item)
        if not os.path.isfile(item_path):
            continue
        if not (item.endswith('.xlsx') or
           item.endswith('.xls')):
            continue
        xlsx_files.append(item_path)
    return xlsx_files

def read_excel_to_dataframes(excel_filepath):
    """Reads an Excel workbook into a dictionary of dataframes keyed by sheet names."""
    dfs = {}
    wb = load_workbook(excel_filepath, data_only=True)
    for sheet_name in wb.sheetnames:
        print('Reading sheet ' + sheet_name)
        # This probably needs an upgraded pandas
        # dfs[sheet_name] = pd.read_excel(xls, sheet_name=sheet_name, engine='xlrd')
        dfs[sheet_name] = pd.read_excel(wb, sheet_name, engine='openpyxl')
    return dfs

def reorder_first_columns(df, first_columns):
    """Reorders a columns in a dataframe so first_columns appear first"""
    return df[move_to_prefix(list(df.columns), first_columns)]

def move_to_prefix(all_list, prefix_list):
    """Reorders elements in a list to move the prefix_list elements first"""
    all_list = list(all_list)  # So we don't mutate all_list
    for p_element in prefix_list:
        all_list.remove(p_element)
    return prefix_list + all_list

def drop_empty_cols(df):
    """Drops columns with empty or null values."""
    df.reset_index(drop=True, inplace=True)
    empty_cols = []
    print(f'Dropping empty values for {df.columns.tolist()}')
    for col in df.columns.tolist():
        indx = ~df[col].isnull()
        if df[indx].empty:
            empty_cols.append(col)
    for col in empty_cols:
        df[col] = np.nan
    df_output = df.dropna(axis=1,how='all').copy()
    df_output.reset_index(drop=True, inplace=True)
    return df_output

def update_multivalue_col_vals(df, multi_col_prefix):
    """Updates the values of multi-value nominal columns"""
    multi_cols = [c for c in df.columns.tolist() if c.startswith(multi_col_prefix)]
    drop_cols = []
    for col in multi_cols:
        df[col] = df[col].astype(str)
        val_index = ((df[col] == '1')|(df[col] == '1.0')|(df[col] == 'True'))
        if df[val_index].empty:
            drop_cols.append(col)
            continue
        # Set rows to the column's value if "True" (1).
        df.loc[val_index, col] = col.split(
            multi_col_prefix
        )[-1].strip()
        # Set rows to blank if the column is not True (1).
        df.loc[~val_index, col] = np.nan
    # Drop the columns that where not actually used.
    df.drop(drop_cols, axis=1, inplace=True, errors='ignore')
    rename_cols = {}
    i = 0
    for col in multi_cols:
        if col in drop_cols:
            continue
        i += 1
        rename_cols[col] = multi_col_prefix + str(i)
    # Rename the columns that were used.
    df.rename(columns=rename_cols, inplace=True)
    return drop_empty_cols(df)

def update_multivalue_columns(df, multival_col_prefixes=None):
    """Updates multivalue columns, removing the ones not in use"""
    if multival_col_prefixes is None:
        multival_col_prefixes = MULTI_VALUE_COL_PREFIXES
    for multi_col_prefix in multival_col_prefixes:
        df = update_multivalue_col_vals(df, multi_col_prefix)
    return df

def clean_up_multivalue_cols(df, skip_cols=[], delim='::'):
    """Cleans up multivalue columns where one column has values that concatenate values from other columns"""
    poss_multi_value_cols = {}
    cols = df.columns.tolist()
    sub_cols = []
    for col in cols:
        if col in skip_cols:
            continue
        for other_col in cols:
            if other_col == col:
                # Same column, skip it.
                continue
            if not other_col.startswith(col) or not '/' in other_col:
                # This other column is not prefixed by col
                continue
            if other_col in sub_cols:
                # We want to avoid a situation where something/1 is considered to be a
                # parent of something/10
                continue
            other_col_vals = df[df[other_col].notnull()][other_col].unique().tolist()
            if len(other_col_vals) > 1:
                # This is not a column with a single value, so skip.
                continue
            sub_cols.append(other_col)
            if col not in poss_multi_value_cols:
                poss_multi_value_cols[col] = []
            # Add a tuple of the other column name, and it's unique value.
            poss_multi_value_cols[col].append((other_col, other_col_vals[0],))
    for col, rel_cols_vals in poss_multi_value_cols.items():
        for act_rel_col, act_val  in rel_cols_vals:
            # Update the col by filtering for non null values for col,
            # and for where the act_rel_col has it's act_val.
            act_val = str(act_val)
            print('Remove the column {} value "{}" in the column {}'.format(act_rel_col, act_val, col))
            rep_indx = (df[col].notnull() & (df[act_rel_col] == act_val))
            # Remove the string we don't want, from col, which concatenates multiple
            # values.
            df.loc[rep_indx, col] = df[col].str.replace(act_val, '')
            if delim in act_val:
                # Remove the first part of a hierarchy delimited value from a likely parent column.
                df.loc[rep_indx, col] = df[col].str.replace(
                    act_val.split(delim)[0],
                    ''
                )
            # Now do final cleanup
            df.loc[rep_indx, col] = df[col].str.strip()
    # Now do a file cleanup, removing anything that's no longer present.
    df = drop_empty_cols(df)
    return df


def split_col_with_delim_into_multiple_cols(df, col, delim=' '):
    """Splits a column with delimited values into multiple columns"""
    df_col = None
    for real_col in df.columns.tolist():
        if col == real_col:
            df_col = real_col
            break
        if real_col.endswith(col):
            df_col = real_col
            break
    if not df_col:
        # we couldn't match the col exactly or as the end of a df columns
        return df
    if not col in df.columns:
        #  Rename the df_col into col to make processing easier
        df.rename(columns={df_col:col}, inplace=True)
    delim_count_col = f'{col}___delim_count'
    df[delim_count_col] = np.nan
    act_index = ~df[col].isnull()
    df.loc[act_index, delim_count_col] = df[act_index][col].str.count(delim)
    max_count = df[delim_count_col].max()
    df.drop(columns=[delim_count_col], inplace=True)
    if max_count < 1:
        # No multiple values, so nothing to split.
        return df
    new_cols = []
    i = 0
    while i <= max_count:
        new_cols.append(f'{col}/{i}')
        i += 1
    df[new_cols] = df[col].str.split(delim, expand=True)
    for act_col in new_cols:
        print(f'Normalize slugs for {act_col}')
        df = make_col_normal_slug_values(df, act_col)
    df.drop(columns=[col], inplace=True)
    df.rename(columns={f'{col}/0': col}, inplace=True)
    df = make_col_normal_slug_values(df, col)
    return df


def split_all_cols_with_delim_into_multiple_cols(
    form_type,
    df,
    config_tups=pc_configs.FORM_COLS_DELIM_SPLIT_TO_MULTIPLE_COLS
):
    """Splits all columns with delimited values into multiple columns"""
    for conf_form_type, col, delim in config_tups:
        if form_type != conf_form_type:
            continue
        df = split_col_with_delim_into_multiple_cols(df, col, delim)
    return df


def get_alternate_labels(
    label,
    project_uuid=pc_configs.PROJECT_UUID,
    config=None
):
    """Returns a list of a label and alternative versions based on project config"""
    label = str(label)
    if config is None:
        config = pc_configs.LABEL_ALTERNATIVE_PARTS
    if not project_uuid in config:
        return [label]
    label_variations = []
    for label_part, label_alts in config[project_uuid].items():
        label_part = str(label_part)
        if not label_part in label:
            label_variations.append(label)
        for label_alt in label_alts:
            label_variations.append(
                label.replace(label_part, label_alt)
            )
    return label_variations

def parse_opencontext_url(s):
    """Returns a tuple of the item_type and uuid for an Open Context url"""
    s = AllManifest().clean_uri(s)
    if not isinstance(s, str):
        return None, None
    if not s.startswith('opencontext.org/'):
        return None, None
    oc_split = s.split('opencontext.org/')
    id_part = oc_split[-1]
    id_parts = id_part.split('/')
    if len(id_parts) < 2:
        # The ID parts is not complete
        return None, None
    item_type = id_parts[0]
    uuid = id_parts[1]
    return item_type, uuid

def parse_opencontext_uuid(s):
    """Returns an Open Context UUID from an OC URL"""
    _, uuid = parse_opencontext_url(s)
    return uuid

def parse_opencontext_type(s):
    """Returns the Open Context item type from an OC URL"""
    item_type, _ = parse_opencontext_url(s)
    return item_type


def get_trench_unit_mapping_dict(trench_id):
    """Gets mapping information for a trench_id based on prefix string"""
    trench_id = str(trench_id)
    for prefix, map_dict in pc_configs.TRENCH_CONTEXT_MAPPINGS.items():
        if not trench_id.startswith(prefix):
            continue
        return map_dict
    return None


def normalize_catalog_label(raw_label):
    """Makes a catalog label fit Poggio Civitate conventions"""
    prefix = ''
    num_part = ''
    for c in str(raw_label):
        if c.isnumeric():
            num_part += c
        else:
            prefix += c
    prefix = prefix.strip()
    return f'{prefix} {num_part}'


def df_fill_in_by_shared_id_cols(df, col_to_fill, id_cols):
    """Fills in the value of a dataframe column based if
    there are values for the same entity identified by values
    in shared ID columns

    :param str col_to_fill: The column with blank values that
        we want to fill with values for the same entities
        identified by values in shared ID columns
    :param list id_cols: A list of columns that are used to
        uniquely identify the same entities

    return df
    """
    grp_cols = [col_to_fill] + id_cols
    if not set(grp_cols).issubset(set(df.columns.tolist())):
        # missing the required columns
        return df
    bad_index = df[col_to_fill].isnull()
    if df[bad_index].empty:
        # Nothing to fill in.
        return df
    good_index = ~df[col_to_fill].isnull()
    if df[good_index].empty:
        # We have no good values to fill by
        return df
    df_g = df[good_index][grp_cols].groupby(grp_cols, as_index=False).first()
    df_g.reset_index(drop=True, inplace=True)
    for _, row in df_g.iterrows():
        # Build an act_index to identify rows with
        # an empty col_to_fill value, but with the same
        # shared id_col values
        act_indx = df[col_to_fill].isnull()
        for c in id_cols:
            act_indx &= (df[c] == row[c])
        df.loc[act_indx, col_to_fill] = row[col_to_fill]
    return df


def get_df_by_sheet_name_part(dfs, sheet_name_part):
    if dfs is None:
        return None, None
    for sheet_name, df in dfs.items():
        if not sheet_name_part in sheet_name:
            continue
        return df, sheet_name
    return None, None


def get_df_with_rel_id_cols(dfs):
    """Gets a dataframe and sheet_name with related ID columns"""
    if dfs is None:
        return None, None
    rel_cols = [c for c,_ in pc_configs.RELS_RENAME_COLS.items()]
    for sheet_name, df in dfs.items():
        if not set(rel_cols).issubset(set(df.columns.tolist())):
            continue
        return df, sheet_name
    return None, None


def prep_df_link_cols(df_link):
    """Prepares columns for a related links dataframe"""
    renames = {c:r_c for c, r_c in pc_configs.RELS_RENAME_COLS.items() if c in df_link.columns}
    df_link.rename(
        columns=renames,
        inplace=True
    )
    check_cols = (
        pc_configs.FIRST_LINK_REL_COLS
        + [
            'object_related_id',
            'object_related_type',
        ]
    )
    for c in check_cols:
        if c in df_link.columns:
            continue
        df_link[c] = np.nan
    return df_link



def get_prepare_df_link_from_rel_id_sheet(dfs):
    """Gets and prepares a df_link from a rel-id sheet"""
    df_link, _ = get_df_with_rel_id_cols(dfs)
    if df_link is None:
        return None
    return prep_df_link_cols(df_link)


def get_general_form_type_from_sheet_name(sheet_name):
    sheet_name = sheet_name.lower()
    form_types = [
        'locus',
        'small find',
        'bulk find',
        'catalog',
        'media',
        'trench book',
    ]
    for act_type in form_types:
        if act_type in sheet_name:
            return act_type
    return None


def get_general_form_type_from_file_sheet_name(file_or_sheet_name):
    file_or_sheet_name = file_or_sheet_name.replace('_', ' ')
    file_or_sheet_name = file_or_sheet_name.replace('-', ' ')
    return get_general_form_type_from_sheet_name(
        file_or_sheet_name
    )


def add_final_subjects_uuid_label_cols(
    df,
    subjects_df,
    form_type,
    final_label_col='subject_label',
    final_uuid_col='subject_uuid',
    final_uuid_source_col='subject_uuid_source',
    orig_uuid_col='_uuid',
):
    """Adds the final (db reconciled) uuids and labels for subjects items"""
    final_cols = [
        final_label_col,
        final_uuid_col,
        final_uuid_source_col,
        'kobo_uuid',
    ]
    for col in final_cols:
        if col in df.columns:
            continue
        df[col] = ''
    if orig_uuid_col not in df.columns:
        return df
    s_label, s_uuid = pc_configs.SUBJECTS_SHEET_PRIMARY_IDs.get(
        form_type,
        (None, None)
    )
    if not s_uuid:
        return df
    if not set([s_label, s_uuid]).issubset(subjects_df.columns.tolist()):
        return df
    for orig_uuid in df[orig_uuid_col].unique().tolist():
        sub_indx = (
            subjects_df['kobo_uuid'] == orig_uuid
        )
        if subjects_df[sub_indx].empty:
            # We don't have this, so can't add these values.
            continue
        df_indx = (
            df[orig_uuid_col] == orig_uuid
        )
        # Update the df final label and final uuid columns with the applicable values
        # from the subjects_df
        df.loc[df_indx, final_label_col] = subjects_df[sub_indx][s_label].iloc[0]
        df.loc[df_indx, final_uuid_col] = subjects_df[sub_indx][s_uuid].iloc[0]
        # Note the kobo provided uuid, which may differ from the uuid we get from
        # the subjects_df
        df.loc[df_indx, 'kobo_uuid'] = orig_uuid
        # Indicate if the uuid as original or if we're using an OC database uuid
        if subjects_df[sub_indx][s_uuid].iloc[0] == orig_uuid:
            df.loc[df_indx, final_uuid_source_col] = pc_configs.UUID_SOURCE_KOBOTOOLBOX
        else:
            df.loc[df_indx, final_uuid_source_col] = pc_configs.UUID_SOURCE_OC_LOOKUP
    return df


def make_trench_supervisor_link_df(
    df,
    person_col='Trench Supervisor',
    link_rel='Supervised by'
):
    """Makes a link dataframe to associate persons """
    cols = ['subject_label', 'subject_uuid', person_col]
    if not set(cols).issubset(set(df.columns.tolist())):
        return None
    df_persons = pd.read_csv(pc_configs.PEOPLE_CSV_PATH)
    df = df[cols].copy()
    df.rename(columns={person_col: 'name'}, inplace=True)
    # This explodes space separated names into multiple rows.
    df_super = df.assign(name=df['name'].str.split(' ')).explode('name').reset_index(drop=True)
    # Merge the df_persons into this.
    df_super = pd.merge(df_super, df_persons, on='name', how='left')
    df_super.rename(
        columns={
            'label_oc': 'object_label',
            'uuid': 'object_uuid',
        },
        inplace=True,
    )
    df_super[pc_configs.LINK_RELATION_TYPE_COL] = link_rel
    final_cols = [
        'subject_label',
        'subject_uuid',
        pc_configs.LINK_RELATION_TYPE_COL,
        'object_label',
        'object_uuid',
        'name',
    ]
    df_super = df_super[final_cols]
    return df_super


def look_up_in_df_persons(lookup_name, df_persons):
    """Return a person label, uuid tuple from the df_persons"""
    lookup_name = lookup_name.strip()
    if not ' ' in lookup_name:
        # Abbreviation only, so we can only use the df_persons for a lookup
        act_index = (df_persons['name'] == lookup_name)
    else:
        name_ex = lookup_name.split(' ')
        act_index = (
            (
                (df_persons['label'] == lookup_name)
            |
                (
                    df_persons['label'].str.startswith(name_ex[0])
                    & df_persons['label'].str.endswith(name_ex[-1])
                )
            )
        )
    if df_persons[act_index].empty:
        return None, None
    return (
        df_persons[act_index]['label_oc'].iloc[0],
        df_persons[act_index]['uuid'].iloc[0],
    )


def add_person_object_rels(
    df,
    person_col,
    link_rel,
    person_label_col='object_label',
    person_uuid_col='object_uuid',
    person_uuid_source_col='object_uuid_source',
):
    cols = [person_col]
    if not set(cols).issubset(set(df.columns.tolist())):
        return df
    add_cols = [
        pc_configs.LINK_RELATION_TYPE_COL,
        person_label_col,
        person_uuid_col,
        person_uuid_source_col,
    ]
    for c in add_cols:
        if c in df.columns:
            continue
        df[c] = np.nan
    df_persons = pd.read_csv(pc_configs.PEOPLE_CSV_PATH)
    act_index = ~df[person_col].isnull()
    for lookup_name in df[act_index][person_col].unique().tolist():
        person_label, person_uuid = look_up_in_df_persons(
            lookup_name,
            df_persons,
        )
        if not person_uuid:
            continue
        up_index = df[person_col] == lookup_name
        df.loc[up_index, pc_configs.LINK_RELATION_TYPE_COL] = link_rel
        df.loc[up_index, person_label_col] = person_label
        df.loc[up_index, person_uuid_col] = person_uuid
        df.loc[up_index, person_uuid_source_col] = 'df_persons'
    return df

def make_col_normal_slug_values(df, col, prefix_for_slugs='24_'):
    """Updates values in a column of a dataframe so slug values are consistent
    with Open Context expectations
    """
    fix_index = (
         ~df[col].isnull()
    )
    df.loc[fix_index, col] = df[fix_index][col].str.strip()
    act_index = (
        ~df[col].isnull()
        & df[col].str.startswith(prefix_for_slugs)
    )
    if df[act_index].empty:
        return df
    df.loc[act_index, col] = df[act_index][col].str.replace('_', '-')
    return df

def make_oc_normal_slug_values(df, prefix_for_slugs='24_'):
    """Updates a dataframe's attribute columns so slug values are consistent
    with Open Context expectations
    """
    for col in df.columns.tolist():
        no_null_index = ~df[col].isnull()
        if not pd.api.types.is_string_dtype(df[no_null_index][col]):
            continue
        df = make_col_normal_slug_values(
            df,
            col,
            prefix_for_slugs=prefix_for_slugs
        )
    return df


def not_null_subject_uuid(df):
    if not 'subject_uuid' in df.columns:
        return df
    act_index = ~df['subject_uuid'].isnull()
    df = df[act_index].copy()
    return df


def get_related_object_labels_from_item_label(item_label):
    """Gets related object labels from a given item label"""
    if not item_label:
        return None
    item_label = str(item_label)
    label_ex = re.split('-|\_|\s', item_label)
    clean_object_labels = []
    has_vdm = False
    has_pc = False
    for label_part in label_ex:
        label_part = label_part.strip()
        if label_part.lower().startswith('vdm'):
            # We have a Vescovado di Murlo object, so make sure we're only
            # using those prefixes
            has_vdm = True
        if label_part.lower().startswith('pc'):
            # We have a PC number so lets make sure we use that prefix
            has_pc = True
    for label_part in label_ex:
        label_part = label_part.strip()
        if len(label_part) < 8:
            continue
        # Get only the number part of the string. This should be a PC/VdM number.
        num_part = ''.join(ch for ch in label_part if ch.isdigit())
        if len(num_part) < 7:
            continue
        if has_pc and not has_vdm:
            clean_object_labels += [f'{prefix}{num_part}' for prefix in pc_configs.PC_LABEL_PREFIXES]
        elif has_vdm and not has_pc:
            clean_object_labels += [f'{prefix}{num_part}' for prefix in pc_configs.VDM_LABEL_PREFIXES]
        else:
            clean_object_labels += [f'{prefix}{num_part}' for prefix in pc_configs.ALL_CATALOG_PREFIXES]
    return clean_object_labels


def get_missing_catalog_item_from_df_subjects(item_label, df_subjects):
    object_label = None
    object_uuid = None
    object_uuid_source = None
    sub_req_cols = ['catalog_name', 'catalog_uuid',]
    if not set(sub_req_cols).issubset(set(df_subjects.columns.tolist())):
        return object_label, object_uuid, object_uuid_source
    clean_object_labels = get_related_object_labels_from_item_label(item_label)
    if not clean_object_labels:
        return object_label, object_uuid, object_uuid_source
    index = df_subjects['catalog_name'].isin(clean_object_labels)
    object_uuids = df_subjects[index]['catalog_uuid'].unique().tolist()
    if len(object_uuids) == 1:
        # we have a good result!
        object_label = df_subjects[index]['catalog_name'].iloc[0]
        object_uuid = df_subjects[index]['catalog_uuid'].iloc[0]
        object_uuid_source = 'subjects.csv'
    return object_label, object_uuid, object_uuid_source


def get_trenchbook_item_from_trench_books_json(
        trench_id,
        year,
        entry_date,
        start_page,
        end_page,
        df_tb,
    ):
    object_label = None
    object_uuid = None
    object_uuid_source = None
    year = int(float(year))
    if year != pc_configs.DEFAULT_IMPORT_YEAR:
        # This will only work for the correct import year
        return object_label, object_uuid, object_uuid_source
    tb_req_cols = ['Trench_ID', 'Date_Documented', 'Start_Page', 'End_Page', 'OC_Label', '_uuid',]
    if not set(tb_req_cols).issubset(set(df_tb.columns.tolist())):
        return object_label, object_uuid, object_uuid_source
    trench_ids = [
        trench_id,
        trench_id.replace('_', '-'),
        trench_id.replace('_', ' '),
        trench_id.lower(),
        trench_id.lower().replace('_', '-'),
        f'{trench_id}_{year}',
        f'{trench_id}_{year}'.lower(),
        f'{trench_id}-{year}',
        f'{trench_id}-{year}'.lower(),
    ]
    if entry_date and not isinstance(entry_date, str):
        entry_date = str(entry_date)
        entry_date = entry_date.split(' ')[0]
    index = (
        (
            df_tb['Trench_ID'].isin(trench_ids)
            & (df_tb['Start_Page'] >= start_page)
            & (df_tb['End_Page'] <= end_page)
            & (df_tb['OC_Label'].str.contains(str(entry_date)))
        )
    )
    object_uuids = df_tb[index]['_uuid'].unique().tolist()
    if len(object_uuids) < 1:
        # Just try to match the entry date, not the page
        index = (
            (
                df_tb['Trench_ID'].isin(trench_ids)
                & (df_tb['OC_Label'].str.contains(str(entry_date)))
            )
        )
        object_uuids = df_tb[index]['_uuid'].unique().tolist()
    if len(object_uuids) < 1:
        return object_label, object_uuid, object_uuid_source
    use_index = index
    if len(object_uuids) > 1:
        new_index = index & (df_tb['Start_Page'] == start_page)
        object_uuids = df_tb[index]['_uuid'].unique().tolist()
        if len(object_uuids) == 1:
            use_index = new_index
        if len(object_uuids) > 1:
            new_index &= (df_tb['End_Page'] == end_page)
            object_uuids = df_tb[index]['_uuid'].unique().tolist()
        if len(object_uuids) == 1:
            use_index = new_index
    # Now use the first result
    object_label = df_tb[use_index]['OC_Label'].iloc[0]
    object_uuid = df_tb[use_index]['_uuid'].iloc[0]
    object_uuid_source = 'trench_books.json'
    return object_label, object_uuid, object_uuid_source


def get_form_data_json(
    form_id,
    token,
    form_url=pc_configs.KOBO_API_URL,
):
    """Gets JSON data from a kobo form"""
    headers = {
        'Authorization': f'Token {token}',
    }
    url = f'{form_url}/api/v2/assets/{form_id}/data/?format=json'
    json_data = None
    try:
        r = requests.get(url, headers=headers)
        r.raise_for_status()
    except:
        return None
    json_data = r.json()
    return json_data


def make_form_data_json_filepath(
    form_type,
    form_id,
    form_year=None,
    save_dir=pc_configs.KOBO_JSON_DATA_PATH,
):
    """Makes a filepath for a form json data file"""
    if form_year:
        file_name = f'{form_type}--{form_year}--{form_id}.json'
    else:
        file_name = f'{form_type}--{form_id}.json'
    file_path = os.path.join(save_dir, file_name)
    return file_path


def get_save_form_data_json(
    form_type,
    form_id,
    token,
    form_year=None,
    form_url=pc_configs.KOBO_API_URL,
    save_dir=pc_configs.KOBO_JSON_DATA_PATH,
):
    """Gets and saves JSON data from a kobo form"""
    json_data = get_form_data_json(
        form_id=form_id,
        token=token,
        form_url=form_url
    )
    if not json_data:
        return None
    file_path = make_form_data_json_filepath(
        form_type=form_type,
        form_id=form_id,
        form_year=form_year,
        save_dir=save_dir,
    )
    with open(file_path, "w") as outfile:
        json.dump(json_data, outfile, indent=4)
    print(f'saved: {file_path}')
    return json_data


def get_save_all_form_data(
    token,
    form_tups=pc_configs.API_FORM_ID_FORM_LABELS_ALL,
    form_url=pc_configs.KOBO_API_URL,
    save_dir=pc_configs.KOBO_JSON_DATA_PATH,
):
    """Iterates through configured forms to fetch json data"""
    for form_id, form_type, form_year in form_tups:
        _ = get_save_form_data_json(
            form_type,
            form_id,
            token,
            form_year=form_year,
            form_url=form_url,
            save_dir=save_dir,
        )


def read_or_fetch_and_save_form_data_json(
    form_type,
    form_id,
    token=None,
    form_year=None,
    form_url=pc_configs.KOBO_API_URL,
    save_dir=pc_configs.KOBO_JSON_DATA_PATH,
):
    json_data = None
    file_path = make_form_data_json_filepath(
        form_type=form_type,
        form_id=form_id,
        form_year=form_year,
        save_dir=save_dir,
    )
    if os.path.exists(file_path):
        with open(file_path, 'r') as openfile:
            # Reading from json file
            json_data = json.load(openfile)
        print(f'Read json data from: {file_path}')
    # if we don't have json_data but we do have a token,
    # try to fetch it from the server
    if not json_data and token:
        json_data = get_save_form_data_json(
            form_type,
            form_id,
            token,
            form_year=form_year,
            form_url=form_url,
            save_dir=save_dir,
        )
    if not json_data and not token:
        raise ValueError(f'Cannot fetch data for {form_type} [{form_id}] provide a Kobo API token for {form_url}')
    return json_data


def read_or_fetch_and_save_form_data__results_json(
    form_type,
    form_id,
    token=None,
    form_year=None,
    form_url=pc_configs.KOBO_API_URL,
    save_dir=pc_configs.KOBO_JSON_DATA_PATH,
    results_dir=pc_configs.KOBO_EXCEL_FILES_PATH,
    results_file_name='trench_books.json',
):
    """Gets """
    results_json_data = None
    results_file_path = os.path.join(results_dir, results_file_name)
    if os.path.exists(results_file_path):
        with open(results_file_path, 'r') as openfile:
            # Reading from json file
            results_json_data = json.load(openfile)
        print(f'Read results_json_data from: {results_file_path}')
    if results_json_data:
        return results_json_data
    # We need the JSON data returned from the API or already previously
    # saved on our file system.
    json_data = read_or_fetch_and_save_form_data_json(
        form_type=form_type,
        form_id=form_id,
        token=token,
        form_year=form_year,
        form_url=form_url,
        save_dir=save_dir,
    )
    results_json_data = json_data.get('results')
    if not results_json_data:
        raise ValueError(f'Cannot fetch data for {form_type} [{form_id}] lacks results')
    with open(results_file_path, "w") as outfile:
        json.dump(results_json_data, outfile, indent=4)
    print(f'Extracted results_json_data from API json_data, saved at {results_file_path}')
    return results_json_data


def get_ids_to_deduplicate(df, label_col, uuid_col):
    delete_suggestions = []
    if not set([label_col, uuid_col]).issubset(set(df.columns.tolist())):
        return delete_suggestions
    index = (
        ~df[label_col].isnull()
        & ~df[uuid_col].isnull()
    )
    label_list = df[index][label_col].unique().tolist()
    for label in label_list:
        act_index = (
            df[label_col] == label
        )
        good_uuid = None
        for _, row in df[act_index].iterrows():
            act_uuid = str(row[uuid_col])
            if good_uuid is None:
                good_uuid = act_uuid
            if act_uuid == good_uuid:
                continue
            delete_sug = {
                'name': label,
                'uuid': act_uuid,
            }
            delete_suggestions.append(delete_sug)
    return delete_suggestions


def redact_suggested_deletions(df, uuid_col, delete_suggestions):
    """Deletes rows by act_label_col and act_uuid_col based on the
    utilities.get_ids_to_deduplicate delete_suggestions list
    """
    if not delete_suggestions:
        return df
    if not set([uuid_col]).issubset(set(df.columns.tolist())):
        # we're missing the columns needed to check for deletions
        return df
    for delete_sug in delete_suggestions:
        if False:
            drop_index = (
                (df[label_col] == delete_sug.get('label'))
                & (df[uuid_col] == delete_sug.get('uuid'))
            )
        drop_index = (
            (df[uuid_col] == delete_sug.get('uuid'))
        )
        df = df[~drop_index].copy()
        df.reset_index(drop=True, inplace=True)
    return df


def remove_col_value_underscores(df, col='Trench'):
    """Removes the underscore from a Trench column"""
    if not col in df.columns.tolist():
        return df
    index = ~df[col].isnull()
    df.loc[index, col] = df[index][col].str.replace('_', ' ')
    return df